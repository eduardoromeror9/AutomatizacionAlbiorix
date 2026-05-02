# from datetime import datetime
# from pathlib import Path
# from openpyxl import Workbook
# from openpyxl.styles import Font


# def export_records_to_excel(records, output_dir: str):
#     Path(output_dir).mkdir(parents=True, exist_ok=True)
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Resultados"

#     headers = [
#         "execution_date", "group_name", "chat_name", "message_date", "message_time",
#         "sender", "detected_name", "detected_rut", "category", "variant",
#         "status", "message_text", "notes"
#     ]
#     ws.append(headers)

#     for cell in ws[1]:
#         cell.font = Font(bold=True)

#     sorted_records = sorted(records, key=lambda r: (r.group_name.lower(), r.message_date, r.message_time, r.category))
#     for r in sorted_records:
#         ws.append([
#             r.execution_date, r.group_name, r.chat_name, r.message_date, r.message_time,
#             r.sender, r.detected_name, r.detected_rut, r.category, r.variant,
#             r.status, r.message_text, r.notes
#         ])

#     ws.freeze_panes = "A2"
#     ws.auto_filter.ref = ws.dimensions

#     widths = {
#         "A": 14, "B": 28, "C": 28, "D": 14, "E": 10,
#         "F": 20, "G": 24, "H": 16, "I": 14, "J": 16,
#         "K": 24, "L": 80, "M": 32
#     }
#     for col, width in widths.items():
#         ws.column_dimensions[col].width = width

#     stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
#     path = Path(output_dir) / f"albiorix_resultados_{stamp}.xlsx"
#     wb.save(path)
#     return str(path)


# from datetime import datetime
# from pathlib import Path
# from openpyxl import Workbook
# from openpyxl.styles import Font, PatternFill


# def export_records_to_excel(records, output_dir: str):
#     Path(output_dir).mkdir(parents=True, exist_ok=True)
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Resultados"

#     headers = [
#         "execution_date", "group_name", "chat_name", "message_date", "message_time",
#         "sender", "detected_name", "detected_rut", "category", "variant",
#         "status", "message_text", "notes"
#     ]
#     ws.append(headers)

#     header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
#     for cell in ws[1]:
#         cell.font = Font(bold=True)
#         cell.fill = header_fill

#     sorted_records = sorted(
#         records,
#         key=lambda r: (
#             r.group_name.lower(),
#             r.message_date,
#             r.message_time,
#             r.detected_name.lower(),
#             r.category.lower()
#         )
#     )

#     for r in sorted_records:
#         ws.append([
#             r.execution_date, r.group_name, r.chat_name, r.message_date, r.message_time,
#             r.sender, r.detected_name, r.detected_rut, r.category, r.variant,
#             r.status, r.message_text, r.notes
#         ])

#     ws.freeze_panes = "A2"
#     ws.auto_filter.ref = ws.dimensions

#     widths = {
#         "A": 14, "B": 20, "C": 20, "D": 14, "E": 10,
#         "F": 20, "G": 28, "H": 16, "I": 14, "J": 16,
#         "K": 24, "L": 60, "M": 30
#     }
#     for col, width in widths.items():
#         ws.column_dimensions[col].width = width

#     stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
#     path = Path(output_dir) / f"albiorix_resultados_{stamp}.xlsx"
#     wb.save(path)
#     return str(path)



# from datetime import datetime
# from pathlib import Path
# from openpyxl import Workbook
# from openpyxl.styles import Font, PatternFill


# def export_records_to_excel(records, output_dir: str):
#     Path(output_dir).mkdir(parents=True, exist_ok=True)
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Resultados"
#     headers = ["execution_date","group_name","chat_name","message_date","message_time","sender","detected_name","detected_rut","category","variant","status","message_text","notes"]
#     ws.append(headers)
#     fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
#     for cell in ws[1]:
#         cell.font = Font(bold=True)
#         cell.fill = fill
#     sorted_records = sorted(records, key=lambda r: (r.group_name.lower(), r.message_date, r.message_time, r.detected_name.lower(), r.category.lower()))
#     for r in sorted_records:
#         ws.append([r.execution_date, r.group_name, r.chat_name, r.message_date, r.message_time, r.sender, r.detected_name, r.detected_rut, r.category, r.variant, r.status, r.message_text, r.notes])
#     ws.freeze_panes = "A2"
#     ws.auto_filter.ref = ws.dimensions
#     widths = {"A":14,"B":20,"C":20,"D":14,"E":10,"F":20,"G":28,"H":16,"I":14,"J":16,"K":24,"L":60,"M":30}
#     for c, w in widths.items():
#         ws.column_dimensions[c].width = w
#     stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
#     path = Path(output_dir) / f"albiorix_resultados_{stamp}.xlsx"
#     wb.save(path)
#     return str(path)

# from datetime import datetime
# from pathlib import Path
# from openpyxl import Workbook
# from openpyxl.styles import Font, PatternFill

# def export_records_to_excel(records, output_dir: str):
#     Path(output_dir).mkdir(parents=True, exist_ok=True)
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Resultados"

#     headers = [
#         "execution_date", "group_name", "chat_name", "message_date", "message_time",
#         "sender", "local", "comuna", "detected_name", "detected_rut",
#         "category", "variant", "status", "message_text", "notes"
#     ]
#     ws.append(headers)

#     fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
#     for cell in ws[1]:
#         cell.font = Font(bold=True)
#         cell.fill = fill

#     sorted_records = sorted(
#         records,
#         key=lambda r: (
#             r.group_name.lower(),
#             r.message_date,
#             r.message_time,
#             r.local.lower(),
#             r.comuna.lower(),
#             r.detected_name.lower(),
#             r.category.lower()
#         )
#     )

#     for r in sorted_records:
#         ws.append([
#             r.execution_date, r.group_name, r.chat_name, r.message_date, r.message_time,
#             r.sender, r.local, r.comuna, r.detected_name, r.detected_rut,
#             r.category, r.variant, r.status, r.message_text, r.notes
#         ])

#     ws.freeze_panes = "A2"
#     ws.auto_filter.ref = ws.dimensions

#     widths = {
#         "A": 14, "B": 20, "C": 20, "D": 14, "E": 10,
#         "F": 20, "G": 18, "H": 20, "I": 24, "J": 16,
#         "K": 14, "L": 16, "M": 14, "N": 58, "O": 28
#     }
#     for col, width in widths.items():
#         ws.column_dimensions[col].width = width

#     stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
#     path = Path(output_dir) / f"albiorix_resultados_{stamp}.xlsx"
#     wb.save(path)
#     return str(path)

# from datetime import datetime
# from pathlib import Path
# from openpyxl import Workbook
# from openpyxl.styles import Font, PatternFill

# def export_records_to_excel(records, output_dir: str):
#     Path(output_dir).mkdir(parents=True, exist_ok=True)
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Resultados"

#     headers = [
#         "execution_date", "group_name", "chat_name", "message_date", "message_time",
#         "sender", "local", "comuna", "detected_name", "detected_rut",
#         "category", "variant", "status", "message_text", "notes"
#     ]
#     ws.append(headers)

#     fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
#     for cell in ws[1]:
#         cell.font = Font(bold=True)
#         cell.fill = fill

#     sorted_records = sorted(
#         records,
#         key=lambda r: (
#             r.group_name.lower(),
#             r.message_date,
#             r.message_time,
#             r.local.lower(),
#             r.comuna.lower(),
#             r.detected_name.lower(),
#             r.category.lower()
#         )
#     )

#     for r in sorted_records:
#         ws.append([
#             r.execution_date, r.group_name, r.chat_name, r.message_date, r.message_time,
#             r.sender, r.local, r.comuna, r.detected_name, r.detected_rut,
#             r.category, r.variant, r.status, r.message_text, r.notes
#         ])

#     ws.freeze_panes = "A2"
#     ws.auto_filter.ref = ws.dimensions

#     widths = {
#         "A": 14, "B": 20, "C": 20, "D": 14, "E": 10,
#         "F": 20, "G": 18, "H": 20, "I": 24, "J": 16,
#         "K": 14, "L": 16, "M": 14, "N": 58, "O": 28
#     }
#     for col, width in widths.items():
#         ws.column_dimensions[col].width = width

#     stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
#     path = Path(output_dir) / f"albiorix_resultados_{stamp}.xlsx"
#     wb.save(path)
#     return str(path)

from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def export_records_to_excel(records, output_dir: str):
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"

    headers = [
        "execution_date", "group_name", "chat_name", "message_date", "message_time",
        "sender", "local", "comuna", "detected_name", "detected_rut",
        "category", "variant", "status", "message_text", "notes"
    ]
    ws.append(headers)

    fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill

    sorted_records = sorted(
        records,
        key=lambda r: (
            (r.group_name or "").lower(),
            (r.message_date or ""),
            (r.message_time or ""),
            (r.local or "").lower(),
            (r.comuna or "").lower(),
            (r.detected_name or "").lower(),
            (r.category or "").lower()
        )
    )

    for r in sorted_records:
        ws.append([
            r.execution_date, r.group_name, r.chat_name, r.message_date, r.message_time,
            r.sender, r.local, r.comuna, r.detected_name, r.detected_rut,
            r.category, r.variant, r.status, r.message_text, r.notes
        ])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = {
        "A": 14, "B": 20, "C": 20, "D": 14, "E": 10,
        "F": 20, "G": 18, "H": 20, "I": 24, "J": 16,
        "K": 14, "L": 16, "M": 14, "N": 58, "O": 28
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = Path(output_dir) / f"albiorix_resultados_{stamp}.xlsx"
    wb.save(path)
    return str(path)